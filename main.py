import requests
import pprint
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
import urllib.request
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.colors import RED, GREEN, YELLOW, BLUE
from openpyxl.cell import Cell


pp = pprint.PrettyPrinter(indent=4)
url_base = "https://pubmed.ncbi.nlm.nih.gov/?"


def main():
    print("Start scraper")
    path = input(
        "Enter excel file name (ex: 'myfile.xlsx' <remove the quotes>): ")

    wb = load_workbook(filename=path)
    ws = wb['Sheet1']

    for row in ws.rows:
        try:
            ref_id = row[0].value
            author = row[1].value
            title = row[2].value
            year = row[3].value
            search_dict = {'term': str(
                author) + ' ' + str(title) + ' ' + str(year)}
            URL = url_base + urlencode(search_dict)
            print(URL)

            page = requests.get(URL)

            soup = BeautifulSoup(page.content, 'html.parser')

            id_list = soup.find(id='full-view-identifiers')
            PMCID = id_list.find('span', class_='identifier pmc')

            if PMCID:
                cell_selection = row[0]
                is_downloaded = run_id_code(PMCID, ref_id, author.split(',')[0], year, cell_selection)
                if is_downloaded:
                    cell_selection.fill = PatternFill(
                        fgColor=GREEN, fill_type='solid'
                    )
            else:
                cell_selection = row[0]

                cell_selection.fill = PatternFill(
                    fgColor=RED, fill_type='solid')
                
                # ! This code can be used to collect the links to individual journals
                # link_div = soup.find('div', class_='full-text-links-list')
                # links = link_div.find_all('a', class_='link-item')

                # download_url = None
                # for link in links:
                #     print("LINK: " + link.get("href").strip())
                #     url = link.get("href").strip()
                #     scraper = JournalScraper()
                #     download_url = scraper.scrape_article(url)
                #     if (download_url is not None):
                #         download_from_url(download_url, ref_id, author, year)
                #         break

                # if download_url is not None:
                #     cell_selection.fill = PatternFill(
                #         fgColor=BLUE, fill_type='solid'
                #     )
                # else:
                #     cell_selection.fill = PatternFill(
                #         fgColor=RED, fill_type='solid'
                #     )

            print("\n")
        except:
            print("Error occurred on the row/webpage, coloured yellow")
            cell_selection = row[0]
            cell_selection.fill = PatternFill(
                fgColor=YELLOW, fill_type='solid')

    wb.save('Updated_Article_list.xlsx')


# def download_from_url(pdf_url, ref_id, author, year):
#     print("Download: " + pdf_url)
#     # response = requests.get(pdf_url)
#     with urllib.request.urlopen(pdf_url) as r:
#         data = r.read()
#         filename = str(ref_id) + '_' + str(author) + '_' + str(year)
#         path = "/Users/Peter/Documents/PDF_Crawler/TestOutput/" + filename + ".pdf"
#         with open(path, 'wb') as f:
#             f.write(data)


def run_id_code(PMCID, ref_id, author, year, cell_selection):
    newURL = "https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id="
    id = PMCID.find('a', class_='id-link').text.strip()
    print("The id is: " + id)
    print("Now run the search with that id")
    print("https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id=" + id)
    request = newURL + id
    new_page = requests.get(request)

    root = ET.fromstring(new_page.content)
    # root = root.find("./records/record")
    if root.findtext("not Open Access"):
        print("Article is not open access on PMC")
        cell_selection.fill = PatternFill(
            fgColor=YELLOW, fill_type='solid'
        )
        return False
    else:
        xml_links = root.findall("./records/record/link")
        for link in xml_links:
            type = link.get('format')
            if type == "pdf":
                pdf_url = link.get("href").strip()
                print("Download: " + pdf_url)
                # response = requests.get(pdf_url)
                with urllib.request.urlopen(pdf_url) as r:
                    data = r.read()
                    filename = str(ref_id) + '_' + \
                        str(author) + '_' + str(year)
                    path = "/Users/Peter/Documents/PDF_Crawler/TestOutput/" + filename + ".pdf"
                    with open(path, 'wb') as f:
                        f.write(data)


if __name__ == '__main__':
    main()
